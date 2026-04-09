#!/usr/bin/env python3
"""
bcic2020_03_preprocess.py

BCIC2020_03 MAT EEG 데이터셋 전처리 → .npy 직접 저장

입력 폴더 구조 예시:
  D:/open_eeg/BCIC2020_03/
    training set/
      Data_Sample01.mat ... Data_Sample15.mat
    validation set/
      Data_Sample01.mat ... Data_Sample15.mat
    test set/
      Data_Sample01.mat ... Data_Sample15.mat

MAT 내부 구조:
  training set   -> epo_train
  validation set -> epo_validation
  test set       -> epo_test

각 struct 내부:
  x : (time, channels, trials)
  y : one-hot, (5, trials) 또는 (trials, 5)
      - 단, test의 y는 비어있을 수 있음 -> Track3_Answer Sheet_Test.xlsx 사용

출력:
  OUT_DIR/
    train/
      eeg.npy, coords.npy, label.npy, meta.json
    val/
      eeg.npy, coords.npy, label.npy, meta.json
    test/
      eeg.npy, coords.npy, label.npy, meta.json

참고:
- 입력 trial이 이미 잘려 있으므로 추가 segmentation은 하지 않음.
- 각 trial은 독립적으로 전처리되며, 필터 경계 영향을 줄이기 위해 reflect padding 후
  filtfilt/sosfiltfilt 내부 padding은 끈다.
"""

import os
import re
import json
import time
import math
import glob
import warnings
from collections import defaultdict

import h5py
import mne
import numpy as np
import openpyxl
import scipy.signal as signal
import scipy.io as sio
from tqdm import tqdm


# ==============================================================================
# [설정 영역]
# ==============================================================================
CONFIG = {
    "ROOT_DIR": r"D:/open_eeg/BCIC2020_03",
    "OUTPUT_DIR": r"D:/open_eeg_eval/BCIC2020_03_npy",
    "ANSWER_SHEET": r"D:/open_eeg/BCIC2020_03/Track3_Answer Sheet_Test.xlsx",

    "TARGET_SR": 200,
    "BANDPASS": (0.5, 75.0),
    "NOTCH_FREQS": 60.0,
    "NOTCH_Q": 30.0,
    "CLIP_LIMIT": 15.0,
    "TRIAL_PAD_SECONDS": 1.0,

    "MONTAGE": "standard_1005",
}


# ==============================================================================
# 1. 유틸
# ==============================================================================
COMMON_64_CH = [
    "Fp1", "Fpz", "Fp2", "AF7", "AF3", "AF4", "AF8", "F7",
    "F5", "F3", "F1", "Fz", "F2", "F4", "F6", "F8",
    "FT7", "FC5", "FC3", "FC1", "FCz", "FC2", "FC4", "FC6",
    "FT8", "T7", "C5", "C3", "C1", "Cz", "C2", "C4",
    "C6", "T8", "TP7", "CP5", "CP3", "CP1", "CPz", "CP2",
    "CP4", "CP6", "TP8", "P7", "P5", "P3", "P1", "Pz",
    "P2", "P4", "P6", "P8", "PO7", "PO3", "POz", "PO4",
    "PO8", "O1", "Oz", "O2", "CB1", "CB2", "HEO", "VEO",
]


def _human_bytes(n: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    x = float(n)
    for u in units:
        if x < 1024.0 or u == units[-1]:
            return f"{x:.2f}{u}"
        x /= 1024.0
    return f"{x:.2f}TB"


def parse_sample_id(file_path: str) -> str:
    m = re.search(r"Data_Sample(\d+)", os.path.basename(file_path), re.IGNORECASE)
    if not m:
        raise ValueError(f"Could not parse sample id from filename: {file_path}")
    return f"Data_Sample{int(m.group(1)):02d}"


def detect_split_and_key(file_path: str):
    parts = [p.lower() for p in os.path.normpath(file_path).split(os.sep)]
    joined = " / ".join(parts)

    if any("training set" in p for p in parts) or "training set" in joined:
        return "train", "epo_train"
    if any("validation set" in p for p in parts) or "validation set" in joined:
        return "val", "epo_validation"
    if any("test set" in p for p in parts) or "test set" in joined:
        return "test", "epo_test"

    raise ValueError(f"Could not infer split from path: {file_path}")


# ==============================================================================
# 2. MATLAB 로딩
# ==============================================================================

def _safe_attr(obj, key, default=None):
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    if hasattr(obj, key):
        return getattr(obj, key)
    return default


def _mat_struct_to_dict(obj):
    """mat_struct / ndarray(object) / dict를 재귀적으로 python object로 변환."""
    if obj is None:
        return None

    if isinstance(obj, dict):
        return {k: _mat_struct_to_dict(v) for k, v in obj.items() if not k.startswith("__")}

    # scipy loadmat의 matlab struct
    if hasattr(obj, "_fieldnames"):
        return {name: _mat_struct_to_dict(getattr(obj, name)) for name in obj._fieldnames}

    if isinstance(obj, np.ndarray):
        if obj.dtype == object:
            if obj.ndim == 0:
                return _mat_struct_to_dict(obj.item())
            return [_mat_struct_to_dict(v) for v in obj.tolist()]
        return obj

    return obj


def _h5_read_any(h5obj, root_file):
    """MAT v7.3(HDF5)용 재귀 loader."""
    if isinstance(h5obj, h5py.Dataset):
        data = h5obj[()]
        # object reference array
        if isinstance(data, np.ndarray) and data.dtype == object:
            out = []
            for ref in data.flatten():
                out.append(_h5_read_any(root_file[ref], root_file))
            return np.array(out, dtype=object).reshape(data.shape)
        return data

    if isinstance(h5obj, h5py.Group):
        return {k: _h5_read_any(v, root_file) for k, v in h5obj.items()}

    return h5obj


def load_epoch_struct(mat_path: str, struct_key: str):
    """scipy.io.loadmat 우선, 실패하면 h5py fallback."""
    # scipy مسیر
    try:
        try:
            mat = sio.loadmat(mat_path, struct_as_record=False, squeeze_me=True, simplify_cells=True)
        except TypeError:
            mat = sio.loadmat(mat_path, struct_as_record=False, squeeze_me=True)
        if struct_key in mat:
            return _mat_struct_to_dict(mat[struct_key])
    except NotImplementedError:
        pass
    except Exception:
        pass

    # HDF5 fallback
    with h5py.File(mat_path, "r") as f:
        if struct_key not in f:
            raise KeyError(f"'{struct_key}' not found in {mat_path}")
        return _h5_read_any(f[struct_key], f)


# ==============================================================================
# 3. 정답 시트 파싱
# ==============================================================================

def load_test_answer_sheet(xlsx_path: str):
    """
    Track3 정답 시트를 읽어
    {"Data_Sample01": np.array([...], dtype=np.int64), ...}
    형태로 반환.

    원본 라벨 1~5 -> 내부 저장용 0~4로 변환.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Track3" not in wb.sheetnames:
        raise KeyError(f"Sheet 'Track3' not found in {xlsx_path}. Found: {wb.sheetnames}")

    ws = wb["Track3"]

    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    mapping = {}

    col = 1
    while col <= ws.max_column:
        title = ws.cell(2, col).value
        next_title = ws.cell(2, col + 1).value if col + 1 <= ws.max_column else None
        next_next = ws.cell(2, col + 2).value if col + 2 <= ws.max_column else None

        # 형태: [Data_SampleXX] [None/blank] ... / row3에서 Trial#, True Label
        if isinstance(title, str) and title.startswith("Data_Sample"):
            trial_col = col
            label_col = col + 1
            labels = []
            for row in range(4, ws.max_row + 1):
                trial_idx = ws.cell(row, trial_col).value
                label = ws.cell(row, label_col).value
                if trial_idx is None or label is None:
                    continue
                labels.append(int(label) - 1)  # 1~5 -> 0~4
            mapping[title] = np.asarray(labels, dtype=np.int64)
            col += 2
        else:
            col += 1

    if not mapping:
        raise ValueError(f"No answer columns parsed from {xlsx_path}")

    return mapping


# ==============================================================================
# 4. 채널 이름 / 좌표
# ==============================================================================

def normalize_channel_name(ch_name: str) -> str:
    if ch_name is None:
        return ""
    clean_name = re.sub(r"[^A-Za-z0-9]", "", str(ch_name)).strip().upper()
    clean_name = re.sub(r"^EEG", "", clean_name)
    name_map = {
        "FP1": "Fp1", "FPZ": "Fpz", "FP2": "Fp2",
        "AFZ": "AFz", "FZ": "Fz", "FCZ": "FCz",
        "CZ": "Cz", "CPZ": "CPz", "PZ": "Pz",
        "POZ": "POz", "OZ": "Oz",
        "T3": "T7", "T4": "T8", "T5": "P7", "T6": "P8",
    }
    return name_map.get(clean_name, clean_name.capitalize())



def _decode_char_codes(arr):
    arr = np.asarray(arr)
    if arr.size == 0:
        return None
    flat = [int(v) for v in arr.reshape(-1).tolist()]
    flat = [v for v in flat if v > 0]
    if not flat:
        return None
    try:
        s = "".join(chr(v) for v in flat if 0 < v < 0x110000)
    except Exception:
        return None
    s = s.strip().replace("\x00", "")
    return s if s else None


def _decode_string_like(value):
    if value is None:
        return None

    if isinstance(value, bytes):
        try:
            value = value.decode("utf-8", errors="ignore")
        except Exception:
            value = value.decode("latin1", errors="ignore")
        value = value.strip()
        return value if value else None

    if isinstance(value, str):
        value = value.strip()
        return value if value else None

    if isinstance(value, (list, tuple)):
        decoded = [_decode_string_like(v) for v in value]
        decoded = [v for v in decoded if v]
        if not decoded:
            return None
        if all(len(v) == 1 for v in decoded):
            return "".join(decoded).strip()
        if len(decoded) == 1:
            return decoded[0]
        return decoded

    if isinstance(value, np.ndarray):
        arr = np.asarray(value)

        # MATLAB char array stored as numeric char codes
        if arr.dtype.kind in ("i", "u"):
            return _decode_char_codes(arr)

        # Native numpy unicode/bytes arrays
        if arr.dtype.kind in ("U", "S"):
            flat = [str(v).strip() for v in arr.reshape(-1).tolist() if str(v).strip()]
            if not flat:
                return None
            if all(len(v) == 1 for v in flat):
                return "".join(flat).strip()
            if len(flat) == 1:
                return flat[0]
            return flat

        # MATLAB cell array / object array
        if arr.dtype == object:
            if arr.ndim == 0:
                return _decode_string_like(arr.item())

            decoded_items = [_decode_string_like(v) for v in arr.reshape(-1).tolist()]
            decoded_items = [v for v in decoded_items if v is not None]
            if not decoded_items:
                return None

            # [ ['F','p','1'], ['F','z'], ... ] or [array([70,112,49]), ...]
            if all(isinstance(v, str) for v in decoded_items):
                return decoded_items

            # flattened nested lists
            flattened = []
            for v in decoded_items:
                if isinstance(v, list):
                    flattened.extend(v)
                else:
                    flattened.append(v)
            flattened = [v for v in flattened if isinstance(v, str) and v]
            return flattened if flattened else None

    # scalar numeric values are not channel names
    return None


def _extract_string_list(candidate, n_channels=None):
    if candidate is None:
        return None

    # Case 1: already a list/tuple/object array of per-channel labels
    decoded = _decode_string_like(candidate)
    if isinstance(decoded, list):
        vals = [str(v).strip() for v in decoded if isinstance(v, str) and str(v).strip()]
        if vals and (n_channels is None or len(vals) == n_channels):
            return vals

    if isinstance(decoded, str):
        # single string is not enough unless n_channels == 1
        if n_channels == 1:
            return [decoded]
        # maybe space-separated labels, uncommon
        parts = [p.strip() for p in re.split(r"[\s,;]+", decoded) if p.strip()]
        if parts and (n_channels is None or len(parts) == n_channels):
            return parts

    # Case 2: MATLAB char matrix where one axis is channels, the other is chars
    if isinstance(candidate, np.ndarray):
        arr = np.asarray(candidate)

        # numeric char codes matrix
        if arr.dtype.kind in ("i", "u") and arr.ndim == 2 and n_channels is not None:
            if arr.shape[0] == n_channels:
                vals = [_decode_char_codes(arr[i, :]) for i in range(arr.shape[0])]
                vals = [v for v in vals if v]
                if len(vals) == n_channels:
                    return vals
            if arr.shape[1] == n_channels:
                vals = [_decode_char_codes(arr[:, i]) for i in range(arr.shape[1])]
                vals = [v for v in vals if v]
                if len(vals) == n_channels:
                    return vals

        # unicode char matrix
        if arr.dtype.kind in ("U", "S") and arr.ndim == 2 and n_channels is not None:
            if arr.shape[0] == n_channels:
                vals = []
                for i in range(arr.shape[0]):
                    s = "".join(str(v) for v in arr[i, :]).strip()
                    vals.append(s)
                vals = [v for v in vals if v]
                if len(vals) == n_channels:
                    return vals
            if arr.shape[1] == n_channels:
                vals = []
                for i in range(arr.shape[1]):
                    s = "".join(str(v) for v in arr[:, i]).strip()
                    vals.append(s)
                vals = [v for v in vals if v]
                if len(vals) == n_channels:
                    return vals

    return None


def extract_channel_names(epoch_struct, n_channels: int):
    candidates = [
        _safe_attr(epoch_struct, "clab"),
        _safe_attr(epoch_struct, "chan_names"),
        _safe_attr(epoch_struct, "channels"),
        _safe_attr(epoch_struct, "channel_names"),
        _safe_attr(epoch_struct, "labels"),
        _safe_attr(epoch_struct, "label"),
    ]

    for cand in candidates:
        vals = _extract_string_list(cand, n_channels=n_channels)
        if vals and len(vals) == n_channels:
            return [normalize_channel_name(v) for v in vals]

    # fallback: 일반적인 64ch 리스트가 길이와 맞으면 사용
    if n_channels == 64:
        return COMMON_64_CH[:]

    return [f"Ch{i+1}" for i in range(n_channels)]



def build_channel_coords(channel_names, montage_name="standard_1005"):
    montage = mne.channels.make_standard_montage(montage_name)
    pos = montage.get_positions()["ch_pos"]

    coords = []
    for ch in channel_names:
        matched = False
        for m_name, m_pos in pos.items():
            if m_name.upper() == ch.upper():
                coords.append(np.asarray(m_pos[:3], dtype=np.float32))
                matched = True
                break
        if not matched:
            coords.append(np.array([0.0, 0.0, 0.0], dtype=np.float32))

    return np.asarray(coords, dtype=np.float16)


# ==============================================================================
# 5. 전처리
# ==============================================================================
class SmartEEGPreprocessor:
    def __init__(self, target_sr, bandpass_freq, clip_limit, notch_freqs, notch_q, trial_pad_seconds):
        self.target_sr = int(target_sr)
        self.bandpass_freq = bandpass_freq
        self.clip_limit = float(clip_limit)
        self.notch_freqs = float(notch_freqs)
        self.notch_q = float(notch_q)
        self.trial_pad_seconds = float(trial_pad_seconds)

    def _pad_reflect(self, eeg_data, original_sr):
        if eeg_data.shape[-1] < 2 or self.trial_pad_seconds <= 0:
            return eeg_data, 0

        pad = int(round(self.trial_pad_seconds * original_sr))
        pad = min(pad, eeg_data.shape[-1] - 1)
        if pad <= 0:
            return eeg_data, 0

        eeg_data = np.pad(eeg_data, ((0, 0), (pad, pad)), mode="reflect")
        return eeg_data, pad

    def apply(self, eeg_data, original_sr):
        """
        eeg_data: (C, T)
        return  : (C, T_resampled)
        """
        eeg_data = np.asarray(eeg_data, dtype=np.float32)
        orig_t = eeg_data.shape[-1]
        expected_t = int(round(orig_t * self.target_sr / float(original_sr)))

        eeg_data, pad = self._pad_reflect(eeg_data, original_sr)
        nyq = 0.5 * float(original_sr)
        low_cut, high_cut = self.bandpass_freq

        adjusted_high = (nyq - 1.0) if high_cut >= nyq else high_cut
        if adjusted_high <= low_cut:
            adjusted_high = min(nyq - 0.1, max(low_cut + 0.1, low_cut * 1.1))

        # 1) multi-notch
        line_freq = self.notch_freqs
        if line_freq and line_freq < nyq:
            b_notch, a_notch = signal.iirnotch(line_freq, Q=self.notch_q, fs=original_sr)
            eeg_data = signal.filtfilt(b_notch, a_notch, eeg_data, axis=-1, padtype=None)

        # 2) bandpass
        wn_low, wn_high = low_cut / nyq, adjusted_high / nyq
        wn_high = min(wn_high, 0.99)
        sos = signal.butter(3, [wn_low, wn_high], btype="band", analog=False, output="sos")
        eeg_data = signal.sosfiltfilt(sos, eeg_data, axis=-1, padtype=None)

        # 3) resample
        if int(round(original_sr)) != self.target_sr:
            gcd = math.gcd(int(round(original_sr)), self.target_sr)
            up = self.target_sr // gcd
            down = int(round(original_sr)) // gcd
            eeg_data = signal.resample_poly(eeg_data, up, down, axis=-1)

        # 4) crop reflect padding in resampled space
        if pad > 0:
            pad_out = int(round(pad * self.target_sr / float(original_sr)))
            if eeg_data.shape[-1] > 2 * pad_out:
                eeg_data = eeg_data[:, pad_out:-pad_out]

        # 5) exact length adjust
        if eeg_data.shape[-1] > expected_t:
            eeg_data = eeg_data[:, :expected_t]
        elif eeg_data.shape[-1] < expected_t:
            pad_right = expected_t - eeg_data.shape[-1]
            eeg_data = np.pad(eeg_data, ((0, 0), (0, pad_right)), mode="edge")

        # 6) z-score + clip
        mean = np.mean(eeg_data, axis=-1, keepdims=True)
        std = np.std(eeg_data, axis=-1, keepdims=True)
        eeg_data = (eeg_data - mean) / (std + 1e-8)
        eeg_data = np.clip(eeg_data, -self.clip_limit, self.clip_limit)
        eeg_data = eeg_data.astype(np.float16)
        return eeg_data[:, :-21]


# ==============================================================================
# 6. x / y 추출
# ==============================================================================

def ensure_numpy_array(x):
    if isinstance(x, np.ndarray):
        return x
    return np.asarray(x)



def _extract_labels(y, split_name: str, sample_id: str, answer_mapping: dict):
    """Return labels as 0-based int64 array.

    - train/val: parse labels from MAT (supports class-index or one-hot)
    - test: always trust the external answer sheet, even if MAT contains a non-empty y
    """
    if split_name == "test":
        if sample_id not in answer_mapping:
            raise KeyError(f"{sample_id} not found in answer sheet")
        return answer_mapping[sample_id].astype(np.int64).reshape(-1)

    labels = None

    if y is not None:
        y_arr = ensure_numpy_array(y)
        if y_arr.size > 0:
            y_arr = np.squeeze(y_arr)
            if y_arr.ndim == 1:
                y_arr = y_arr.astype(np.int64).reshape(-1)
                if y_arr.size == 0:
                    labels = None
                elif y_arr.min() >= 1 and y_arr.max() <= 5:
                    labels = y_arr - 1
                else:
                    labels = y_arr
            elif y_arr.ndim == 2:
                if 5 in y_arr.shape:
                    # one-hot: (5, N) -> argmax(axis=0), (N, 5) -> argmax(axis=1)
                    cls_axis = 0 if y_arr.shape[0] == 5 else 1
                    labels = np.argmax(y_arr, axis=cls_axis).astype(np.int64).reshape(-1)
                else:
                    raise ValueError(
                        f"Unsupported 2D label shape in {sample_id}: y.shape={y_arr.shape}"
                    )
            else:
                raise ValueError(
                    f"Unsupported label ndim in {sample_id}: y.ndim={y_arr.ndim}, y.shape={y_arr.shape}"
                )

    if labels is None:
        raise ValueError(f"Missing/empty labels in non-test split for {sample_id}")

    return labels



def _infer_x_axes(x: np.ndarray, labels: np.ndarray, sample_id: str):
    """
    Infer axes of x robustly.
    Target output order is (trials, channels, time).

    We use:
    - channel axis: preferably size 64
    - trial axis: axis matching len(labels) if unique
    - time axis: remaining axis, preferably the largest of the rest

    This is needed because scipy.loadmat and h5py fallback can expose MAT arrays
    in different dimension orders.
    """
    shape = tuple(int(v) for v in x.shape)
    expected_trials = int(labels.shape[0])

    channel_candidates = [i for i, s in enumerate(shape) if s == 64]
    if len(channel_candidates) == 1:
        channel_axis = channel_candidates[0]
    elif len(channel_candidates) > 1:
        # Prefer the middle axis when ambiguous.
        channel_axis = 1 if 1 in channel_candidates else channel_candidates[0]
    else:
        # Fallback: choose the axis that is neither the largest (usually time)
        # nor the label-matching axis (usually trials), if possible.
        trial_match = [i for i, s in enumerate(shape) if s == expected_trials]
        if len(trial_match) == 1:
            remaining = [i for i in range(3) if i != trial_match[0]]
            channel_axis = min(remaining, key=lambda i: shape[i])
        else:
            channel_axis = min(range(3), key=lambda i: abs(shape[i] - 64))

    trial_candidates = [i for i, s in enumerate(shape) if s == expected_trials and i != channel_axis]
    if len(trial_candidates) == 1:
        trial_axis = trial_candidates[0]
    elif len(trial_candidates) > 1:
        # Prefer the first/last non-channel axis depending on common layouts.
        trial_axis = trial_candidates[0]
    else:
        # Fallback: among non-channel axes, the smaller one is usually trials.
        non_channel_axes = [i for i in range(3) if i != channel_axis]
        trial_axis = min(non_channel_axes, key=lambda i: shape[i])

    time_axis = [i for i in range(3) if i not in (trial_axis, channel_axis)]
    if len(time_axis) != 1:
        raise ValueError(f"Failed to infer x axes for {sample_id}: x.shape={shape}")
    time_axis = time_axis[0]

    return trial_axis, channel_axis, time_axis



def extract_xy(epoch_struct, split_name: str, sample_id: str, answer_mapping: dict):
    x = _safe_attr(epoch_struct, "x")
    y = _safe_attr(epoch_struct, "y")

    if x is None:
        raise KeyError(f"'x' not found in {sample_id} ({split_name})")

    x = ensure_numpy_array(x)
    if x.ndim != 3:
        raise ValueError(f"Expected x to be 3D, got shape={x.shape} in {sample_id}")

    labels = _extract_labels(y, split_name, sample_id, answer_mapping)
    trial_axis, channel_axis, time_axis = _infer_x_axes(x, labels, sample_id)

    eeg = np.transpose(x, (trial_axis, channel_axis, time_axis)).astype(np.float32)  # (N, C, T)
    n_trials = eeg.shape[0]

    if labels.shape[0] != n_trials:
        raise ValueError(
            f"Label count mismatch in {sample_id}: labels={labels.shape[0]}, n_trials={n_trials}, "
            f"x.shape={x.shape}, inferred_axes=(trial={trial_axis}, channel={channel_axis}, time={time_axis})"
        )

    return eeg, labels


# ==============================================================================
# 7. 파일 처리
# ==============================================================================

def process_single_file(file_path: str, answer_mapping: dict, preprocessor: SmartEEGPreprocessor):
    split_name, struct_key = detect_split_and_key(file_path)
    sample_id = parse_sample_id(file_path)
    epoch_struct = load_epoch_struct(file_path, struct_key)

    raw_x = _safe_attr(epoch_struct, "x")
    raw_x_shape = tuple(int(v) for v in ensure_numpy_array(raw_x).shape) if raw_x is not None else None

    eeg_trials, labels = extract_xy(epoch_struct, split_name, sample_id, answer_mapping)
    n_trials, n_channels, _ = eeg_trials.shape

    # channel names / coords
    channel_names = extract_channel_names(epoch_struct, n_channels)
    coords = build_channel_coords(channel_names, CONFIG["MONTAGE"])
    coords_trials = np.repeat(coords[None, :, :], n_trials, axis=0)

    # preprocess each trial independently
    processed_trials = []
    for trial_idx in range(n_trials):
        processed = preprocessor.apply(eeg_trials[trial_idx], original_sr=256)
        processed_trials.append(processed)
    eeg_out = np.stack(processed_trials, axis=0)  # (N, C, T)

    meta_items = []
    for trial_idx in range(n_trials):
        meta_items.append({
            "source_file": os.path.relpath(file_path, CONFIG["ROOT_DIR"]),
            "sample_id": sample_id,
            "split": split_name,
            "trial_idx": int(trial_idx),
            "label": int(labels[trial_idx]),
            "label_original": int(labels[trial_idx] + 1),
            "n_channels": int(n_channels),
            "channel_names": channel_names,
            "original_sfreq": 256,
            "target_sfreq": CONFIG["TARGET_SR"],
            "raw_x_shape": list(raw_x_shape) if raw_x_shape is not None else None,
        })

    return split_name, eeg_out, coords_trials.astype(np.float16), labels.astype(np.int8), meta_items


# ==============================================================================
# 8. 저장
# ==============================================================================

def save_split(split_name: str, split_dir: str, eeg_list, coords_list, label_list, meta_list, file_count: int):
    if not eeg_list:
        print(f"[{split_name}] No valid trials. Skipping.")
        return 0

    os.makedirs(split_dir, exist_ok=True)

    eeg_arr = np.concatenate(eeg_list, axis=0)
    coords_arr = np.concatenate(coords_list, axis=0)
    label_arr = np.concatenate(label_list, axis=0)

    eeg_path = os.path.join(split_dir, "eeg.npy")
    coords_path = os.path.join(split_dir, "coords.npy")
    label_path = os.path.join(split_dir, "label.npy")
    meta_path = os.path.join(split_dir, "meta.json")

    np.save(eeg_path, eeg_arr)
    np.save(coords_path, coords_arr)
    np.save(label_path, label_arr)

    unique, counts = np.unique(label_arr, return_counts=True)
    label_distribution = {int(k): int(v) for k, v in zip(unique, counts)}

    meta = {
        "source": CONFIG["ROOT_DIR"],
        "split": split_name,
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "num_trials": int(eeg_arr.shape[0]),
        "num_files": int(file_count),
        "arrays": {
            "eeg": {"shape": list(eeg_arr.shape), "dtype": str(eeg_arr.dtype)},
            "coords": {"shape": list(coords_arr.shape), "dtype": str(coords_arr.dtype)},
            "label": {"shape": list(label_arr.shape), "dtype": str(label_arr.dtype)},
        },
        "files": {
            "eeg": {"path": eeg_path, "size": _human_bytes(os.path.getsize(eeg_path))},
            "coords": {"path": coords_path, "size": _human_bytes(os.path.getsize(coords_path))},
            "label": {"path": label_path, "size": _human_bytes(os.path.getsize(label_path))},
        },
        "config": {
            "target_sr": CONFIG["TARGET_SR"],
            "bandpass": list(CONFIG["BANDPASS"]),
            "notch_freqs": CONFIG["NOTCH_FREQS"],
            "notch_q": CONFIG["NOTCH_Q"],
            "clip_limit": CONFIG["CLIP_LIMIT"],
            "trial_pad_seconds": CONFIG["TRIAL_PAD_SECONDS"],
        },
        "label_mapping": {
            "original_1": 0,
            "original_2": 1,
            "original_3": 2,
            "original_4": 3,
            "original_5": 4,
        },
        "label_distribution": label_distribution,
        "items": meta_list,
    }

    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    dist_str = ", ".join(f"{k}:{v}" for k, v in label_distribution.items())
    print(f"[{split_name}] Saved {eeg_arr.shape[0]} trials ({file_count} files)")
    print(f"  - eeg   : {_human_bytes(os.path.getsize(eeg_path))}  shape={eeg_arr.shape}")
    print(f"  - coords: {_human_bytes(os.path.getsize(coords_path))}  shape={coords_arr.shape}")
    print(f"  - label : {_human_bytes(os.path.getsize(label_path))}  shape={label_arr.shape}")
    print(f"  - label dist (0~4): {dist_str}\n")

    return int(eeg_arr.shape[0])


# ==============================================================================
# 9. 메인
# ==============================================================================

def collect_files(root_dir: str):
    files = sorted(glob.glob(os.path.join(root_dir, "**", "Data_Sample*.mat"), recursive=True))
    split_files = defaultdict(list)
    for f in files:
        split_name, _ = detect_split_and_key(f)
        split_files[split_name].append(f)
    return split_files


if __name__ == "__main__":
    os.makedirs(CONFIG["OUTPUT_DIR"], exist_ok=True)

    print(f"[init] ROOT_DIR     : {CONFIG['ROOT_DIR']}")
    print(f"[init] OUTPUT_DIR   : {CONFIG['OUTPUT_DIR']}")
    print(f"[init] ANSWER_SHEET : {CONFIG['ANSWER_SHEET']}")
    print(f"[init] TARGET_SR    : {CONFIG['TARGET_SR']}")
    print(f"[init] BANDPASS     : {CONFIG['BANDPASS']}")
    print(f"[init] NOTCH_FREQS  : {CONFIG['NOTCH_FREQS']}")
    print(f"[init] trial pad    : {CONFIG['TRIAL_PAD_SECONDS']}s\n")

    answer_mapping = load_test_answer_sheet(CONFIG["ANSWER_SHEET"])
    print(f"[init] Loaded answer sheet for {len(answer_mapping)} test files")

    split_files = collect_files(CONFIG["ROOT_DIR"])
    print(f"[split] Train: {len(split_files['train'])} files")
    print(f"[split] Val  : {len(split_files['val'])} files")
    print(f"[split] Test : {len(split_files['test'])} files\n")

    preprocessor = SmartEEGPreprocessor(
        target_sr=CONFIG["TARGET_SR"],
        bandpass_freq=CONFIG["BANDPASS"],
        clip_limit=CONFIG["CLIP_LIMIT"],
        notch_freqs=CONFIG["NOTCH_FREQS"],
        notch_q=CONFIG["NOTCH_Q"],
        trial_pad_seconds=CONFIG["TRIAL_PAD_SECONDS"],
    )

    buckets = {
        "train": {"eeg": [], "coords": [], "label": [], "meta": [], "file_count": 0},
        "val": {"eeg": [], "coords": [], "label": [], "meta": [], "file_count": 0},
        "test": {"eeg": [], "coords": [], "label": [], "meta": [], "file_count": 0},
    }

    all_files = split_files["train"] + split_files["val"] + split_files["test"]
    for file_path in tqdm(all_files, desc="[process]"):
        try:
            split_name, eeg_arr, coords_arr, label_arr, meta_items = process_single_file(
                file_path, answer_mapping, preprocessor
            )
            buckets[split_name]["eeg"].append(eeg_arr)
            buckets[split_name]["coords"].append(coords_arr)
            buckets[split_name]["label"].append(label_arr)
            buckets[split_name]["meta"].extend(meta_items)
            buckets[split_name]["file_count"] += 1
        except Exception as e:
            print(f"[ERROR] {file_path}: {e}")

    total = 0
    total += save_split(
        "train",
        os.path.join(CONFIG["OUTPUT_DIR"], "train"),
        buckets["train"]["eeg"],
        buckets["train"]["coords"],
        buckets["train"]["label"],
        buckets["train"]["meta"],
        buckets["train"]["file_count"],
    )
    total += save_split(
        "val",
        os.path.join(CONFIG["OUTPUT_DIR"], "val"),
        buckets["val"]["eeg"],
        buckets["val"]["coords"],
        buckets["val"]["label"],
        buckets["val"]["meta"],
        buckets["val"]["file_count"],
    )
    total += save_split(
        "test",
        os.path.join(CONFIG["OUTPUT_DIR"], "test"),
        buckets["test"]["eeg"],
        buckets["test"]["coords"],
        buckets["test"]["label"],
        buckets["test"]["meta"],
        buckets["test"]["file_count"],
    )

    print("=" * 60)
    print(f"[Done] All splits processed. Total trials: {total}")
    print("=" * 60)
